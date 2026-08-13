#!/usr/bin/env python3
"""Merge Lean 4 ledger workbooks and log every run into the atom registry.

Preservation-first policy:
- The newest mission-control workbook is used as the base.
- Older or duplicate workbooks are fingerprinted, not deleted.
- Addendum sheets are copied with an ADDENDUM_ prefix.
- Every run writes a runtime receipt, a markdown report, a JSONL run log, and
  an atom-ledger event.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REPO = Path(__file__).resolve().parents[1]
MASTER_EXCEL = Path(r"\\192.168.2.50\h_hp\Desktop\Master EXCEL")
DEFAULT_OUTPUT = MASTER_EXCEL / "Lean 4 - CANONICAL_LEDGER_MERGED.xlsx"
DEFAULT_RUNTIME = REPO / "_runtime" / "lean4_ledger_runs"
DEFAULT_ATOM = REPO / "_ledger" / "atoms" / "tp-lean4-canonical-ledger.json"
DEFAULT_REPORT = REPO / "_ledger" / "LEAN4_CANONICAL_LEDGER_MERGE_REPORT.md"
DEFAULT_JSONL = REPO / "_ledger" / "LEAN4_RUN_LOG.jsonl"

BASE_WORKBOOK = Path(r"\\192.168.2.50\h_hp\Desktop\Documents\LEAN 4 Master.xlsx")
SOURCE_WORKBOOKS = [
    Path(r"\\192.168.2.50\h_hp\Desktop\Master EXCEL\Lean 4 - CANONICAL_LEDGER_V2.xlsx"),
    Path(r"\\192.168.2.50\h_hp\Desktop\Master EXCEL\Lean 4 - CANONICAL_LEDGER_V2 - Python Colab Audit.xlsx"),
    Path(r"\\192.168.2.50\h_hp\Desktop\Master EXCEL\Lean 4 - CANONICAL_LEDGER_V2.backup_before_python_colab_audit.xlsx"),
    Path(r"\\192.168.2.50\h_hp\Desktop\Master EXCEL\Lean 4 - CANONICAL_LEDGER_V2.pre_python_colab_replace.xlsx"),
    Path(r"\\192.168.2.50\h_hp\Desktop\Master EXCEL\Lean 4.backup_20260710_041459.xlsx"),
    BASE_WORKBOOK,
    Path(r"\\192.168.2.50\h_hp\Desktop 2\Theophysics_Lean4_Addendum_Updated (1).xlsx"),
]
ADDENDUM_WORKBOOK = Path(r"\\192.168.2.50\h_hp\Desktop 2\Theophysics_Lean4_Addendum_Updated (1).xlsx")

GENERATED_SHEETS = {
    "MERGE_README",
    "MERGE_SOURCE_MANIFEST",
    "MERGE_DUPLICATE_GROUPS",
}


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def stable_sha(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="ignore")).hexdigest()


def file_sha(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def sheet_stats(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        out = []
        for ws in wb.worksheets:
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            out.append({"name": ws.title, "rows": rows, "cols": cols})
        return out
    finally:
        wb.close()


def source_manifest(paths: list[Path]) -> list[dict[str, Any]]:
    rows = []
    for path in paths:
        item: dict[str, Any] = {
            "path": str(path),
            "exists": path.exists(),
            "role": "source",
        }
        if path.exists():
            stat = path.stat()
            item.update(
                {
                    "bytes": stat.st_size,
                    "modified": datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(timespec="seconds"),
                    "sha256": file_sha(path),
                    "sheets": sheet_stats(path),
                }
            )
        rows.append(item)
    return rows


def duplicate_groups(manifest: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[str]] = {}
    for item in manifest:
        sha = item.get("sha256")
        if sha:
            grouped.setdefault(sha, []).append(item["path"])
    return [{"sha256": sha, "paths": paths} for sha, paths in grouped.items() if len(paths) > 1]


def write_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(80, max(12, col_idx * 2 + 10))


def copy_values_to_sheet(src_ws, dest_ws) -> None:
    for row in src_ws.iter_rows(values_only=True):
        dest_ws.append(list(row))


def remove_generated_sheets(wb) -> None:
    for name in list(wb.sheetnames):
        if name in GENERATED_SHEETS or name.startswith("ADDENDUM_"):
            del wb[name]


def build_workbook(manifest: list[dict[str, Any]], output: Path, run_id: str) -> dict[str, Any]:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(BASE_WORKBOOK)
    remove_generated_sheets(wb)

    readme = wb.create_sheet("MERGE_README", 0)
    write_rows(
        readme,
        [
            ["Lean 4 Canonical Ledger Merge"],
            ["run_id", run_id],
            ["generated_at", now_iso()],
            ["canonical_policy", "LEAN 4 Master.xlsx is base; V2 workbooks fingerprinted; addendum sheets copied with ADDENDUM_ prefix."],
            ["cleanup_policy", "Do not delete source workbooks until this merged workbook is reviewed."],
            ["atom_log", str(DEFAULT_ATOM)],
        ],
    )

    source_ws = wb.create_sheet("MERGE_SOURCE_MANIFEST")
    write_rows(
        source_ws,
        [["path", "exists", "bytes", "modified", "sha256", "sheet_count", "sheet_names"]]
        + [
            [
                item["path"],
                item["exists"],
                item.get("bytes", ""),
                item.get("modified", ""),
                item.get("sha256", ""),
                len(item.get("sheets", [])),
                ", ".join(sheet["name"] for sheet in item.get("sheets", [])),
            ]
            for item in manifest
        ],
    )

    dup_ws = wb.create_sheet("MERGE_DUPLICATE_GROUPS")
    dup_rows = [["sha256", "duplicate_paths"]]
    for group in duplicate_groups(manifest):
        dup_rows.append([group["sha256"], "\n".join(group["paths"])])
    write_rows(dup_ws, dup_rows)

    addendum_count = 0
    if ADDENDUM_WORKBOOK.exists():
        add_wb = load_workbook(ADDENDUM_WORKBOOK, read_only=True, data_only=True)
        try:
            for src_ws in add_wb.worksheets:
                title = ("ADDENDUM_" + src_ws.title)[:31]
                if title in wb.sheetnames:
                    del wb[title]
                dest_ws = wb.create_sheet(title)
                copy_values_to_sheet(src_ws, dest_ws)
                addendum_count += 1
        finally:
            add_wb.close()

    wb.save(output)
    return {
        "output_path": str(output),
        "output_sha256": file_sha(output),
        "addendum_sheets_added": addendum_count,
        "sheet_count": len(wb.sheetnames),
    }


def atom_uid(payload: dict[str, Any]) -> str:
    basis = json.dumps(
        {
            "atom_id": payload["atom_id"],
            "title": payload["title"],
            "claim": payload["claim"],
            "domain": payload["domain"],
            "lane": payload["lane"],
        },
        sort_keys=True,
    )
    return "sha256:" + stable_sha(basis)


def event_id(event: dict[str, Any]) -> str:
    return "sha256:" + stable_sha(json.dumps(event, sort_keys=True, default=str))


def upsert_atom(run: dict[str, Any], atom_path: Path) -> dict[str, Any]:
    atom_path.parent.mkdir(parents=True, exist_ok=True)
    if atom_path.exists():
        atom = json.loads(atom_path.read_text(encoding="utf-8"))
    else:
        atom = {
            "atom_id": "tp:lane4/lean4/canonical-ledger",
            "title": "Lean 4 Canonical Ledger",
            "claim": "Lean 4 ledger workbooks require a single canonical merged ledger with provenance and run receipts.",
            "domain": "lean4",
            "lane": "lean4",
            "claim_class": "registry_control",
            "mode_classification": "active_registry",
            "assumptions": [
                "Excel workbooks are administrative ledgers, not Lean kernel proofs.",
                "Source workbooks must be preserved until merge review is complete.",
            ],
            "definitions": ["Canonical Lean ledger merge control atom."],
            "equations": [],
            "bridges": [],
            "dependencies": [],
            "negative_guards": ["Do not treat Excel merge status as proof status."],
            "kill_conditions": ["A reviewed canonical ledger is established elsewhere and supersedes this control atom."],
            "proof_label": "RERUN_OWED",
            "current_status": "active_candidate",
            "rerun_status": "required_on_ledger_change",
            "source_artifacts": [],
            "ledger": [],
        }
        atom["atom_uid"] = atom_uid(atom)

    event = {
        "timestamp": run["generated_at"],
        "lane": "lean4",
        "event_type": "lean4_ledger_merge_run",
        "result": "merged",
        "artifact_path": run["workbook"]["output_path"],
        "meaning": f"Merged Lean 4 ledger workbooks; added {run['workbook']['addendum_sheets_added']} addendum sheets.",
        "limits": "Workbook merge is provenance/registry hygiene; it does not certify Lean theorem truth.",
        "reviewer": "merge_lean4_ledgers.py",
        "event_uuid": str(uuid.uuid4()),
    }
    event["event_id"] = event_id(event)
    atom["ledger"].append(event)
    for item in run["sources"]:
        if item.get("exists"):
            atom["source_artifacts"].append(item["path"])
    atom["source_artifacts"].append(run["workbook"]["output_path"])
    atom["source_artifacts"] = sorted(set(atom["source_artifacts"]))
    atom_path.write_text(json.dumps(atom, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return atom


def write_report(run: dict[str, Any], report_path: Path) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    dupes = duplicate_groups(run["sources"])
    lines = [
        "# Lean 4 Canonical Ledger Merge Report",
        "",
        f"- Run ID: `{run['run_id']}`",
        f"- Generated: `{run['generated_at']}`",
        f"- Output workbook: `{run['workbook']['output_path']}`",
        f"- Output SHA-256: `{run['workbook']['output_sha256']}`",
        f"- Source workbook count: {len(run['sources'])}",
        f"- Duplicate source groups: {len(dupes)}",
        f"- Addendum sheets added: {run['workbook']['addendum_sheets_added']}",
        f"- Atom log: `{run['atom_path']}`",
        "",
        "## Source Workbooks",
        "",
        "| Path | Exists | Bytes | SHA-256 | Sheets |",
        "|---|---:|---:|---|---:|",
    ]
    for item in run["sources"]:
        lines.append(
            f"| `{item['path']}` | {item['exists']} | {item.get('bytes', '')} | `{item.get('sha256', '')}` | {len(item.get('sheets', []))} |"
        )
    lines.extend(["", "## Duplicate Groups", ""])
    if dupes:
        for group in dupes:
            lines.append(f"- `{group['sha256']}`")
            for path in group["paths"]:
                lines.append(f"  - `{path}`")
    else:
        lines.append("- none")
    lines.extend(
        [
            "",
            "## Cleanup Recommendation",
            "",
            "Keep `LEAN 4 Master.xlsx` and the merged workbook as active review surfaces.",
            "Archive duplicate exact copies only after human review of this report.",
            "Do not delete any source workbook; move retired copies into a dated archive folder with this report as receipt.",
        ]
    )
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def append_jsonl(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False, sort_keys=True) + "\n")


def main() -> int:
    parser = argparse.ArgumentParser(description="Merge Lean 4 ledger workbooks and log to atom registry.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--runtime-dir", type=Path, default=DEFAULT_RUNTIME)
    parser.add_argument("--atom", type=Path, default=DEFAULT_ATOM)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    args = parser.parse_args()

    run_id = "lean4-ledger-merge-" + datetime.now().strftime("%Y%m%d-%H%M%S")
    manifest = source_manifest(SOURCE_WORKBOOKS)
    workbook_result = build_workbook(manifest, args.output, run_id)
    run = {
        "run_id": run_id,
        "generated_at": now_iso(),
        "sources": manifest,
        "workbook": workbook_result,
        "atom_path": str(args.atom),
        "report_path": str(args.report),
    }
    args.runtime_dir.mkdir(parents=True, exist_ok=True)
    runtime_path = args.runtime_dir / f"{run_id}.json"
    runtime_path.write_text(json.dumps(run, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    run["runtime_path"] = str(runtime_path)
    upsert_atom(run, args.atom)
    write_report(run, args.report)
    append_jsonl(DEFAULT_JSONL, run)
    print(json.dumps({"run_id": run_id, "output": str(args.output), "runtime": str(runtime_path)}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
